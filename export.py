import json

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

from config import ANALYSIS_JSON, ANALYSIS_XLSX
from exam import load_results
from questions import load_questions


def export_analysis_to_excel():
    if not ANALYSIS_JSON.exists():
        raise FileNotFoundError("No analysis.json exists yet. Extract item analysis first.")

    try:
        analysis = json.loads(ANALYSIS_JSON.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise ValueError("analysis.json is invalid.") from exc

    question_map = {q["id"]: q for q in load_questions()}

    item_rows = []
    for item in analysis.get("items", []):
        question = question_map.get(item["question_id"], {})
        wrong_students = item.get("students_wrong", [])
        wrong_text = "; ".join(
            f"{student['student_name']} ({student['student_answer']})"
            for student in wrong_students
        ) if wrong_students else "None"
        item_rows.append({
            "Question ID": item["question_id"],
            "Question": question.get("question", item.get("question", "")),
            "Module": question.get("module", item.get("module", "")),
            "Question Type": question.get("type", item.get("question_type", "")),
            "Correct Answer": question.get("answer", item.get("correct_answer", "")),
            "Total Responses": item.get("total_responses", 0),
            "Number Correct": item.get("number_correct", 0),
            "Number Wrong": item.get("number_wrong", 0),
            "Difficulty Index": item.get("difficulty_index"),
            "Discrimination Index": item.get("discrimination_index"),
            "Recommendation": item.get("recommendation", "Review"),
            "Students Wrong": wrong_text,
        })

    summary = pd.DataFrame([
        ["Exam", analysis["exam"]],
        ["Number of Students", analysis["students"]],
        ["Number of Questions", analysis["questions"]],
        ["Reliability", analysis["reliability"] if analysis["reliability"] is not None else "N/A"],
    ], columns=["Summary", "Value"])

    student_rows = []
    for student in analysis.get("student_results", []):
        student_rows.append({
            "Student Name": student.get("student_name", "Unknown"),
            "Exam": analysis.get("exam", ""),
            "Score": student.get("score", 0),
            "Total Questions": student.get("total_questions", 0),
            "Percentage": student.get("percentage", 0.0),
            "Timestamp": student.get("timestamp", ""),
        })

    response_rows = []
    for response in load_results():
        if response.get("exam_id") != analysis.get("exam_id"):
            continue
        question = question_map.get(response.get("question_id"), {})
        response_rows.append({
            "Question ID": response.get("question_id", ""),
            "Question": question.get("question", ""),
            "Student Name": response.get("student_name", ""),
            "Student Answer": response.get("student_answer", ""),
            "Correct Answer": response.get("correct_answer", ""),
            "Is Correct": response.get("is_correct", False),
            "Student Score": response.get("score", 0),
            "Student Percentage": response.get("percentage", 0.0),
            "Timestamp": response.get("timestamp", ""),
        })

    items_df = pd.DataFrame(item_rows)
    student_df = pd.DataFrame(student_rows)
    response_df = pd.DataFrame(response_rows)

    with pd.ExcelWriter(ANALYSIS_XLSX, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        items_df.to_excel(writer, sheet_name="Item Analysis", index=False)
        student_df.to_excel(writer, sheet_name="Student Results", index=False)
        response_df.to_excel(writer, sheet_name="Question Responses", index=False)

    workbook = load_workbook(ANALYSIS_XLSX)
    for worksheet in workbook.worksheets:
        if worksheet.max_row == 1:
            continue
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
        for column_cells in worksheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(max_length + 2, 60)
    workbook.save(ANALYSIS_XLSX)

    return ANALYSIS_XLSX
